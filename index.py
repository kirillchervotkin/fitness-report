# -*- coding: utf-8 -*-
import os
import re
import threading
import unicodedata
import shutil
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment

# ========================= НАСТРОЙКИ ПО УМОЛЧАНИЮ =========================
DEFAULT_SETTINGS = {
    "week_filter_mode": "latest",
    "target_year": 2026,
    "target_week": 12,
    "league_filter": "",
    "days_green": 5,
    "days_yellow": 4,
    "dur_green": "05:00:00",
    "dur_yellow_low": "04:30:00",
    "cardio_green": 300,
    "cardio_yellow_low": 250,
    "output_base": "polar_training_report",
    "fill_green_hex": "C6EFCE",
    "fill_yellow_hex": "FFEB9C",
    "fill_red_hex": "FFC7CE",
}

# ========================= ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ =========================
def normalize_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()

def normalize_columns(columns):
    return [str(col).strip() for col in columns]

def detect_column(df, keywords, required=False, label="column"):
    normalized_map = {str(col).strip().lower(): col for col in df.columns}
    for keyword in keywords:
        keyword = keyword.lower().strip()
        for norm_col, orig_col in normalized_map.items():
            if norm_col == keyword:
                return orig_col
    for keyword in keywords:
        keyword = keyword.lower().strip()
        for norm_col, orig_col in normalized_map.items():
            if keyword in norm_col:
                return orig_col
    if required:
        raise ValueError(f"Required {label} not found. Columns: {list(df.columns)}")
    return None

def normalize_merge_key(value):
    text = normalize_text(value)
    text = re.sub(r"\s+", "", text)
    text = text.replace("_", "")
    text = unicodedata.normalize('NFKD', text)
    text = ''.join(c for c in text if not unicodedata.combining(c))
    return text.casefold()

def read_csv_robust(file_path):
    best_df = None
    best_score = -10**9
    for sep in [";", ","]:
        for encoding in ["utf-8-sig", "utf-8", "cp1251", "latin1"]:
            try:
                df = pd.read_csv(file_path, sep=sep, encoding=encoding)
                df.columns = normalize_columns(df.columns)
                cols = [str(c).lower() for c in df.columns]
                score = len(df.columns)
                if len(df.columns) == 1:
                    score -= 1000
                expected_groups = [
                    ["имя", "name", "athlete", "user"],
                    ["день", "date", "day"],
                    ["продолжительность", "duration"],
                    ["кардионагрузка", "cardio load", "тренировочная нагрузка"],
                ]
                for group in expected_groups:
                    if any(any(k in col for k in group) for col in cols):
                        score += 100
                if score > best_score:
                    best_score = score
                    best_df = df.copy()
            except Exception:
                pass
    if best_df is None:
        raise ValueError(f"Could not read CSV file: {file_path}")
    return best_df

def parse_duration_to_seconds(series):
    numeric_seconds = pd.to_numeric(series, errors='coerce')
    timedelta_seconds = pd.to_timedelta(series.astype(str).str.strip(), errors='coerce').dt.total_seconds()
    result = numeric_seconds.combine_first(timedelta_seconds)
    return result

def apply_role_group(value):
    text = normalize_text(value).lower()
    assistant_keywords = ["assistant", "assist", "ассистент", "помощник", "лайнсмен"]
    referee_keywords = ["referee", "judge", "судья", "главный"]
    if any(k in text for k in assistant_keywords):
        return "Assistant"
    if any(k in text for k in referee_keywords):
        return "Referee"
    return "Other"

def safe_sheet_name(name, used_names):
    name = normalize_text(name)
    if not name:
        name = "Blank"
    name = re.sub(r'[:\\/?*\[\]]', "_", name)
    name = name[:31].strip()
    if not name:
        name = "Sheet"
    original_name = name
    counter = 1
    while name in used_names:
        suffix = f"_{counter}"
        name = (original_name[:31 - len(suffix)] + suffix).strip()
        counter += 1
    used_names.add(name)
    return name

def format_workbook(output_path, settings):
    wb = load_workbook(output_path)
    green_fill = PatternFill(fill_type="solid", start_color=settings["fill_green_hex"], end_color=settings["fill_green_hex"])
    yellow_fill = PatternFill(fill_type="solid", start_color=settings["fill_yellow_hex"], end_color=settings["fill_yellow_hex"])
    red_fill = PatternFill(fill_type="solid", start_color=settings["fill_red_hex"], end_color=settings["fill_red_hex"])
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(fill_type="solid", start_color="D9EAF7", end_color="D9EAF7")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    dur_green_seconds = pd.to_timedelta(settings["dur_green"]).total_seconds()
    dur_yellow_seconds = pd.to_timedelta(settings["dur_yellow_low"]).total_seconds()
    
    for ws in wb.worksheets:
        headers = {}
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value is not None:
                headers[cell.value] = idx
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Применяем рамки ко всем ячейкам в используемом диапазоне
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border
        
        if "Total Duration" in headers:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row_idx, headers["Total Duration"]).number_format = "[h]:mm:ss"
        
        for date_col_name in ["Week Start", "Week End"]:
            if date_col_name in headers:
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row_idx, headers[date_col_name]).number_format = "dd.mm.yyyy"
        
        required_cols = ["Training Days", "Total Duration", "Total Duration Seconds", "Total Cardioload"]
        if all(col in headers for col in required_cols):
            for row_idx in range(2, ws.max_row + 1):
                days_cell = ws.cell(row_idx, headers["Training Days"])
                dur_cell = ws.cell(row_idx, headers["Total Duration"])
                dur_seconds_cell = ws.cell(row_idx, headers["Total Duration Seconds"])
                cardio_cell = ws.cell(row_idx, headers["Total Cardioload"])
                
                days_val = days_cell.value if days_cell.value is not None else 0
                cardio_val = cardio_cell.value if cardio_cell.value is not None else 0
                try:
                    dur_seconds = float(dur_seconds_cell.value) if dur_seconds_cell.value is not None else 0.0
                except Exception:
                    dur_seconds = 0.0
                
                if days_val >= settings["days_green"]:
                    days_cell.fill = green_fill
                elif days_val == settings["days_yellow"]:
                    days_cell.fill = yellow_fill
                else:
                    days_cell.fill = red_fill
                
                if dur_seconds > dur_green_seconds:
                    dur_cell.fill = green_fill
                elif dur_yellow_seconds <= dur_seconds <= dur_green_seconds:
                    dur_cell.fill = yellow_fill
                else:
                    dur_cell.fill = red_fill
                
                if cardio_val > settings["cardio_green"]:
                    cardio_cell.fill = green_fill
                elif settings["cardio_yellow_low"] <= cardio_val <= settings["cardio_green"]:
                    cardio_cell.fill = yellow_fill
                else:
                    cardio_cell.fill = red_fill
        
        if "Total Duration Seconds" in headers:
            col_letter = ws.cell(row=1, column=headers["Total Duration Seconds"]).column_letter
            ws.column_dimensions[col_letter].hidden = True
        
        for col_cells in ws.columns:
            max_len = 0
            letter = col_cells[0].column_letter
            for cell in col_cells:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 40)
        
        ws.freeze_panes = "A2"
    
    wb.save(output_path)

# ========================= ОСНОВНАЯ ЛОГИКА =========================
def generate_report(polar_file_paths, master_file_path, settings, log_callback):
    def log(msg):
        if log_callback:
            log_callback(msg)

    import os
    import tempfile
    import zipfile
    import glob
    import pandas as pd

    temp_dir = tempfile.mkdtemp(prefix="polar_extract_")
    csv_paths = []
    for path in polar_file_paths:
        if path.lower().endswith(".zip"):
            with zipfile.ZipFile(path, "r") as zf:
                zf.extractall(temp_dir)
        elif path.lower().endswith(".csv"):
            csv_paths.append(path)
    csv_paths.extend(glob.glob(os.path.join(temp_dir, "**", "*.csv"), recursive=True))
    csv_paths = sorted(list(dict.fromkeys(csv_paths)))
    if not csv_paths:
        raise ValueError("No CSV files found in selected Polar data.")

    log(f"Найдено CSV файлов: {len(csv_paths)}")

    frames = []
    for path in csv_paths:
        df = read_csv_robust(path)
        df.columns = normalize_columns(df.columns)

        df['source_file'] = os.path.splitext(os.path.basename(path))[0]

        # Определяем тип файла
        is_polar = any(col in df.columns for col in ['Sport', 'Вид спорта'])
        # Для надёжности: если есть start_date_local и нет Sport, то это Garmin
        if not is_polar and 'start_date_local' in df.columns:
            is_polar = False
        elif 'Sport' in df.columns:
            is_polar = True

        if is_polar:
            date_col = detect_column(df, ["день", "дата", "date", "day", "start_date_local"], required=True, label="date")
            duration_col = detect_column(df, ["продолжительность", "длительность", "duration", "moving_time"], required=True, label="duration")
            cardio_col = detect_column(df, ["кардионагрузка", "тренировочная нагрузка", "cardio load", "hr_load"], required=True, label="cardio load")
            user_col = detect_column(df, ["имя", "спортсмен", "athlete", "user", "name"], required=False)
            athlete_name = df[user_col] if user_col is not None else df['source_file']
        else:
            date_col = detect_column(df, ["start_date_local"], required=True, label="date")
            duration_col = detect_column(df, ["moving_time"], required=True, label="duration")
            cardio_col = detect_column(df, ["hr_load"], required=True, label="cardio load")
            athlete_name = df['source_file'].str.split('_').str[0]

        date_series = df[date_col].astype(str).str.strip()

        if is_polar:
            # Polar: формат день-месяц-год, разделитель '-' или '.'
            sample = date_series.dropna().iloc[0] if not date_series.dropna().empty else ""
            if '.' in sample:
                sep = '.'
            else:
                sep = '-'
            if ':' in sample:
                session_dt = pd.to_datetime(date_series, format=f'%d{sep}%m{sep}%Y %H:%M:%S', errors='coerce')
            else:
                session_dt = pd.to_datetime(date_series, format=f'%d{sep}%m{sep}%Y', errors='coerce')
        else:
            # Garmin: стандартный ISO (год-месяц-день)
            session_dt = pd.to_datetime(date_series, errors='coerce')
            log(f"Garmin дата образец: '{date_series.iloc[0]}' -> '{session_dt.iloc[0]}'")

        duration_sec = parse_duration_to_seconds(df[duration_col]).fillna(0)
        cardio_load = pd.to_numeric(df[cardio_col], errors='coerce').fillna(0)

        unified_df = pd.DataFrame({
            'athlete_name': athlete_name,
            'Session DateTime': session_dt,
            'Duration Seconds': duration_sec,
            'Total Cardioload': cardio_load,
            'source_file': df['source_file']
        })
        unified_df = unified_df[unified_df['Session DateTime'].notna()].copy()
        frames.append(unified_df)

    polar_df = pd.concat(frames, ignore_index=True)

    polar_df["CSV Name"] = polar_df["athlete_name"].map(normalize_text)
    polar_df["Session Date"] = polar_df["Session DateTime"].dt.normalize()

    polar_df = polar_df[(polar_df["CSV Name"] != "") & (polar_df["Session DateTime"].notna())].copy()
    if polar_df.empty:
        raise ValueError("No valid Polar rows after cleaning.")

    iso = polar_df["Session DateTime"].dt.isocalendar()
    polar_df["Year"] = iso["year"].astype(int)
    polar_df["Week Number"] = iso["week"].astype(int)

    if settings["week_filter_mode"] == "latest":
        latest = polar_df[["Year", "Week Number"]].drop_duplicates().sort_values(["Year", "Week Number"]).tail(1)
        selected_year = int(latest.iloc[0]["Year"])
        selected_week = int(latest.iloc[0]["Week Number"])
    else:
        selected_year = settings["target_year"]
        selected_week = settings["target_week"]

    log(f"Выбрана неделя: {selected_year} / {selected_week}")
    polar_df = polar_df[(polar_df["Year"] == selected_year) & (polar_df["Week Number"] == selected_week)].copy()
    if polar_df.empty:
        raise ValueError("No data for selected week.")

    before = len(polar_df)
    polar_df = polar_df.drop_duplicates(subset=["CSV Name", "Session Date", "Duration Seconds", "Total Cardioload"], keep="first")
    log(f"Удалено дубликатов строк: {before - len(polar_df)}")

    aggregated = (polar_df.groupby(["CSV Name", "Year", "Week Number"])
                  .agg(Training_Days=("Session Date", "nunique"),
                       Total_Duration_Seconds=("Duration Seconds", "sum"),
                       Total_Cardioload=("Total Cardioload", "sum"))
                  .reset_index())
    aggregated["Training_Days"] = aggregated["Training_Days"].astype(int)

    aggregated.rename(columns={
        "Training_Days": "Training Days",
        "Total_Duration_Seconds": "Total Duration Seconds",
        "Total_Cardioload": "Total Cardioload"
    }, inplace=True)

    log(f"Агрегировано записей: {len(aggregated)}")

    use_master = master_file_path and os.path.exists(master_file_path)
    master_effective = False
    
    if use_master:
        xls = pd.ExcelFile(master_file_path)
        frames_master = []
        for sheet in xls.sheet_names:
            df_sheet = xls.parse(sheet)
            df_sheet.columns = normalize_columns(df_sheet.columns)
            frames_master.append(df_sheet)
        master_df = pd.concat(frames_master, ignore_index=True)
    
        referee_col = detect_column(master_df, ["referee name", "имя", "full name", "name"], required=True)
        polar_col = detect_column(master_df, ["polar name", "имя polar", "polar"], required=True)
        gender_col = detect_column(master_df, ["gender", "пол"], required=True)
        role_col = detect_column(master_df, ["role", "роль"], required=True)
        league_col = detect_column(master_df, ["league", "лига"], required=True)
        master_df = master_df.rename(columns={referee_col: "Referee Name", polar_col: "Polar Name",
                                              gender_col: "Gender", role_col: "Role", league_col: "League"})
        for col in ["Referee Name", "Polar Name", "Gender", "Role", "League"]:
            master_df[col] = master_df[col].fillna("").map(normalize_text)
        master_df = master_df[master_df["Referee Name"] != ""].copy().drop_duplicates(subset=["Referee Name"])
        
        league_filter_norm = normalize_text(settings["league_filter"])
        if league_filter_norm:
            master_df = master_df[master_df["League"] == league_filter_norm].copy()
        
        if not master_df.empty:
            master_effective = True
            log(f"Мастер-лист загружен, строк: {len(master_df)}")
            aggregated["_csv_key"] = aggregated["CSV Name"].map(normalize_merge_key)
            master_df["_polar_key"] = master_df["Polar Name"].map(normalize_merge_key)
            master_df["_ref_key"] = master_df["Referee Name"].map(normalize_merge_key)
            csv_keys = set(aggregated["_csv_key"].dropna())
            polar_overlap = len(csv_keys & set(master_df["_polar_key"].dropna()))
            ref_overlap = len(csv_keys & set(master_df["_ref_key"].dropna()))
            merge_mode = "Polar Name" if polar_overlap >= ref_overlap else "Referee Name"
            merge_col = "_polar_key" if merge_mode == "Polar Name" else "_ref_key"
            log(f"Режим слияния: {merge_mode} (совпадений {polar_overlap} vs {ref_overlap})")
            final_df = master_df.merge(aggregated, left_on=merge_col, right_on="_csv_key", how="left", suffixes=("", "_agg"))
            final_df["Year"] = final_df["Year"].fillna(selected_year).astype(int)
            final_df["Week Number"] = final_df["Week Number"].fillna(selected_week).astype(int)
            final_df["Training Days"] = final_df["Training Days"].fillna(0).astype(int)
            final_df["Total Duration Seconds"] = final_df["Total Duration Seconds"].fillna(0.0)
            final_df["Total Cardioload"] = final_df["Total Cardioload"].fillna(0.0)
            final_df.drop(columns=["_csv_key", "_polar_key", "_ref_key"], errors="ignore", inplace=True)
            if "CSV Name" in final_df.columns:
                final_df.drop(columns=["CSV Name"], inplace=True)
            final_df["Comments"] = ""  # Пустое поле для комментариев
        else:
            log("Мастер-лист пуст после фильтрации по лиге. Отчёт будет построен без мастер-листа.")
    
    if not master_effective:
        final_df = aggregated.copy()
        final_df["Referee Name"] = final_df["CSV Name"]
        final_df["Gender"] = ""
        final_df["Role"] = ""
        final_df["League"] = ""
        final_df["Comments"] = ""  # Пустое поле для комментариев
        final_df.drop(columns=["CSV Name"], inplace=True)

    week_start = pd.Timestamp.fromisocalendar(selected_year, selected_week, 1)
    week_end = pd.Timestamp.fromisocalendar(selected_year, selected_week, 7)
    final_df["Week Start"] = week_start
    final_df["Week End"] = week_end
    final_df["Total Duration"] = pd.to_timedelta(final_df["Total Duration Seconds"], unit="s")
    final_df["Total Cardioload"] = final_df["Total Cardioload"].round(2)

    for col in ["Referee Name", "Gender", "Role", "League"]:
        if col in final_df.columns:
            final_df[col] = final_df[col].fillna("").map(normalize_text)

    required_final_cols = ["Referee Name", "Gender", "Role", "League", "Year", "Week Number",
                           "Week Start", "Week End", "Training Days", "Total Duration",
                           "Total Duration Seconds", "Total Cardioload", "Comments"]
    for col in required_final_cols:
        if col not in final_df.columns:
            if col in ["Training Days", "Year", "Week Number"]:
                final_df[col] = 0
            elif col in ["Total Duration Seconds", "Total Cardioload"]:
                final_df[col] = 0.0
            elif col == "Total Duration":
                final_df[col] = pd.Timedelta(0)
            elif col == "Comments":
                final_df[col] = ""
            else:
                final_df[col] = ""

    final_df = final_df[required_final_cols].copy()

    extra_sheets = {}
    if master_effective:
        role_groups = final_df["Role"].apply(apply_role_group)
        referees = final_df[role_groups == "Referee"].copy()
        assistants = final_df[role_groups == "Assistant"].copy()
        if not referees.empty:
            extra_sheets["Referees"] = referees
        if not assistants.empty:
            extra_sheets["Assistants"] = assistants
        for league in final_df["League"].dropna().unique():
            league_norm = normalize_text(league)
            if not league_norm:
                continue
            league_df = final_df[final_df["League"] == league_norm].copy()
            extra_sheets[f"League_{league_norm}"] = league_df
            league_ref = league_df[league_df["Role"].apply(apply_role_group) == "Referee"].copy()
            league_asst = league_df[league_df["Role"].apply(apply_role_group) == "Assistant"].copy()
            if not league_ref.empty:
                extra_sheets[f"{league_norm}_Referees"] = league_ref
            if not league_asst.empty:
                extra_sheets[f"{league_norm}_Assistants"] = league_asst

    league_suffix = f"_{normalize_text(settings['league_filter'])}" if settings["league_filter"] else ""
    week_start_str = week_start.strftime("%d.%m.%Y")
    week_end_str = week_end.strftime("%d.%m.%Y")
    output_filename = (f"{settings['output_base']}{league_suffix}_{selected_year}_week{selected_week}_"
                       f"{week_start_str}-{week_end_str}.xlsx")
    output_path = os.path.join(os.getcwd(), output_filename)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        final_df.to_excel(writer, sheet_name="All", index=False)
        used = {"All"}
        for name, df_sheet in extra_sheets.items():
            safe_name = safe_sheet_name(name, used)
            df_sheet.to_excel(writer, sheet_name=safe_name, index=False)
    format_workbook(output_path, settings)

    shutil.rmtree(temp_dir, ignore_errors=True)
    return output_path

# ========================= ГРАФИЧЕСКИЙ ИНТЕРФЕЙС =========================
class PolarReportApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Polar Training Report Generator")
        self.root.geometry("800x700")

        self.polar_files = []
        self.master_file = None

        self.settings = DEFAULT_SETTINGS.copy()
        self.week_filter_var = tk.StringVar(value=self.settings["week_filter_mode"])
        self.target_year_var = tk.StringVar(value=str(self.settings["target_year"]))
        self.target_week_var = tk.StringVar(value=str(self.settings["target_week"]))
        self.league_filter_var = tk.StringVar(value=self.settings["league_filter"])
        self.days_green_var = tk.StringVar(value=str(self.settings["days_green"]))
        self.days_yellow_var = tk.StringVar(value=str(self.settings["days_yellow"]))
        self.dur_green_var = tk.StringVar(value=self.settings["dur_green"])
        self.dur_yellow_low_var = tk.StringVar(value=self.settings["dur_yellow_low"])
        self.cardio_green_var = tk.StringVar(value=str(self.settings["cardio_green"]))
        self.cardio_yellow_low_var = tk.StringVar(value=str(self.settings["cardio_yellow_low"]))
        self.output_base_var = tk.StringVar(value=self.settings["output_base"])

        self.create_widgets()

    def create_widgets(self):
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        file_frame = ttk.LabelFrame(main_frame, text="1. Данные Polar (CSV / ZIP)", padding="5")
        file_frame.pack(fill=tk.X, pady=5)

        self.polar_listbox = tk.Listbox(file_frame, height=4)
        self.polar_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0,5))
        scroll_polar = ttk.Scrollbar(file_frame, orient=tk.VERTICAL, command=self.polar_listbox.yview)
        scroll_polar.pack(side=tk.RIGHT, fill=tk.Y)
        self.polar_listbox.config(yscrollcommand=scroll_polar.set)

        btn_frame = ttk.Frame(file_frame)
        btn_frame.pack(side=tk.RIGHT, fill=tk.Y)
        ttk.Button(btn_frame, text="Добавить файлы", command=self.add_polar_files).pack(pady=2)
        ttk.Button(btn_frame, text="Очистить список", command=self.clear_polar_files).pack(pady=2)

        master_frame = ttk.LabelFrame(main_frame, text="2. Мастер-лист (Excel) – опционально", padding="5")
        master_frame.pack(fill=tk.X, pady=5)

        self.master_label = ttk.Label(master_frame, text="Не выбран")
        self.master_label.pack(side=tk.LEFT, padx=5)
        ttk.Button(master_frame, text="Выбрать файл", command=self.select_master_file).pack(side=tk.RIGHT, padx=5)
        ttk.Button(master_frame, text="Очистить", command=self.clear_master_file).pack(side=tk.RIGHT, padx=5)

        settings_frame = ttk.LabelFrame(main_frame, text="3. Параметры отчёта", padding="5")
        settings_frame.pack(fill=tk.X, pady=5)

        week_mode_frame = ttk.Frame(settings_frame)
        week_mode_frame.pack(fill=tk.X, pady=2)
        ttk.Label(week_mode_frame, text="Выбор недели:").pack(side=tk.LEFT)
        ttk.Radiobutton(week_mode_frame, text="Последняя доступная", variable=self.week_filter_var, value="latest").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(week_mode_frame, text="Вручную:", variable=self.week_filter_var, value="manual").pack(side=tk.LEFT, padx=5)
        self.year_spin = ttk.Spinbox(week_mode_frame, from_=2020, to=2030, width=6, textvariable=self.target_year_var)
        self.year_spin.pack(side=tk.LEFT, padx=2)
        ttk.Label(week_mode_frame, text="год").pack(side=tk.LEFT)
        self.week_spin = ttk.Spinbox(week_mode_frame, from_=1, to=53, width=4, textvariable=self.target_week_var)
        self.week_spin.pack(side=tk.LEFT, padx=2)
        ttk.Label(week_mode_frame, text="неделя").pack(side=tk.LEFT)

        league_frame = ttk.Frame(settings_frame)
        league_frame.pack(fill=tk.X, pady=2)
        ttk.Label(league_frame, text="Фильтр по лиге (оставьте пустым для всех):").pack(side=tk.LEFT)
        ttk.Entry(league_frame, textvariable=self.league_filter_var, width=20).pack(side=tk.LEFT, padx=5)

        thresh_frame = ttk.LabelFrame(settings_frame, text="Пороговые значения", padding="5")
        thresh_frame.pack(fill=tk.X, pady=5)

        grid = ttk.Frame(thresh_frame)
        grid.pack()
        ttk.Label(grid, text="Training Days:").grid(row=0, column=0, sticky=tk.W, padx=5)
        ttk.Label(grid, text="Зелёный ≥").grid(row=0, column=1)
        ttk.Entry(grid, width=5, textvariable=self.days_green_var).grid(row=0, column=2)
        ttk.Label(grid, text="Жёлтый =").grid(row=0, column=3, padx=5)
        ttk.Entry(grid, width=5, textvariable=self.days_yellow_var).grid(row=0, column=4)

        ttk.Label(grid, text="Продолжительность:").grid(row=1, column=0, sticky=tk.W, padx=5)
        ttk.Label(grid, text="Зелёный >").grid(row=1, column=1)
        ttk.Entry(grid, width=8, textvariable=self.dur_green_var).grid(row=1, column=2)
        ttk.Label(grid, text="Жёлтый от").grid(row=1, column=3, padx=5)
        ttk.Entry(grid, width=8, textvariable=self.dur_yellow_low_var).grid(row=1, column=4)

        ttk.Label(grid, text="Cardio load:").grid(row=2, column=0, sticky=tk.W, padx=5)
        ttk.Label(grid, text="Зелёный >").grid(row=2, column=1)
        ttk.Entry(grid, width=5, textvariable=self.cardio_green_var).grid(row=2, column=2)
        ttk.Label(grid, text="Жёлтый от").grid(row=2, column=3, padx=5)
        ttk.Entry(grid, width=5, textvariable=self.cardio_yellow_low_var).grid(row=2, column=4)

        out_frame = ttk.Frame(settings_frame)
        out_frame.pack(fill=tk.X, pady=2)
        ttk.Label(out_frame, text="База имени выходного файла:").pack(side=tk.LEFT)
        ttk.Entry(out_frame, textvariable=self.output_base_var, width=30).pack(side=tk.LEFT, padx=5)

        btn_frame2 = ttk.Frame(main_frame)
        btn_frame2.pack(fill=tk.X, pady=10)
        self.run_btn = ttk.Button(btn_frame2, text="Сформировать отчёт", command=self.run_report)
        self.run_btn.pack()

        log_frame = ttk.LabelFrame(main_frame, text="Лог выполнения", padding="5")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        self.log_text = scrolledtext.ScrolledText(log_frame, height=12, state=tk.NORMAL)
        self.log_text.pack(fill=tk.BOTH, expand=True)

    def add_polar_files(self):
        files = filedialog.askopenfilenames(
            title="Выберите CSV или ZIP файлы Polar",
            filetypes=[("CSV files", "*.csv"), ("ZIP files", "*.zip"), ("All files", "*.*")]
        )
        for f in files:
            if f not in self.polar_files:
                self.polar_files.append(f)
                self.polar_listbox.insert(tk.END, os.path.basename(f))
        self.log(f"Добавлено файлов Polar: {len(files)}")

    def clear_polar_files(self):
        self.polar_files = []
        self.polar_listbox.delete(0, tk.END)
        self.log("Список Polar файлов очищен")

    def select_master_file(self):
        path = filedialog.askopenfilename(title="Выберите Excel файл мастер-листа", filetypes=[("Excel files", "*.xlsx")])
        if path:
            self.master_file = path
            self.master_label.config(text=os.path.basename(path))
            self.log(f"Выбран мастер-лист: {path}")

    def clear_master_file(self):
        self.master_file = None
        self.master_label.config(text="Не выбран")
        self.log("Мастер-лист отключён")

    def log(self, message):
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def update_settings_from_gui(self):
        self.settings["week_filter_mode"] = self.week_filter_var.get()
        try:
            self.settings["target_year"] = int(self.target_year_var.get())
            self.settings["target_week"] = int(self.target_week_var.get())
        except ValueError:
            raise ValueError("Год и неделя должны быть числами")
        self.settings["league_filter"] = self.league_filter_var.get().strip()
        self.settings["days_green"] = int(self.days_green_var.get())
        self.settings["days_yellow"] = int(self.days_yellow_var.get())
        self.settings["dur_green"] = self.dur_green_var.get().strip()
        self.settings["dur_yellow_low"] = self.dur_yellow_low_var.get().strip()
        self.settings["cardio_green"] = int(self.cardio_green_var.get())
        self.settings["cardio_yellow_low"] = int(self.cardio_yellow_low_var.get())
        self.settings["output_base"] = self.output_base_var.get().strip()

    def run_report(self):
        if not self.polar_files:
            messagebox.showerror("Ошибка", "Не выбраны файлы Polar данных")
            return
        try:
            self.update_settings_from_gui()
        except Exception as e:
            messagebox.showerror("Ошибка в настройках", str(e))
            return

        self.run_btn.config(state=tk.DISABLED, text="Обработка...")
        self.log("Начало формирования отчёта...")

        def task():
            try:
                output_path = generate_report(
                    self.polar_files,
                    self.master_file,
                    self.settings,
                    self.log
                )
                save_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx")],
                    initialfile=os.path.basename(output_path),
                    title="Сохранить отчёт как"
                )
                if save_path:
                    shutil.move(output_path, save_path)
                    self.log(f"Отчёт сохранён: {save_path}")
                    messagebox.showinfo("Готово", f"Отчёт успешно создан:\n{save_path}")
                else:
                    self.log("Сохранение отменено пользователем. Временный файл: " + output_path)
            except Exception as e:
                self.log(f"ОШИБКА: {e}")
                messagebox.showerror("Ошибка", str(e))
            finally:
                self.root.after(0, lambda: self.run_btn.config(state=tk.NORMAL, text="Сформировать отчёт"))

        threading.Thread(target=task, daemon=True).start()

if __name__ == "__main__":
    root = tk.Tk()
    app = PolarReportApp(root)
    root.mainloop()